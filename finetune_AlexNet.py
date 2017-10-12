"""
Ref: https://kratzert.github.io/2017/02/24/finetuning-alexnet-with-tensorflow.html

Author: Frederik Kratzert 
Edit: Huiting Hong
"""
import os
import numpy as np
import tensorflow as tf
from datetime import datetime
from alexnet import AlexNet
from simplenn import SimpleNN
from datagenerator import ImageDataGenerator
import pandas as pd
from openpyxl import load_workbook

# Path to the textfiles for the trainings and validation set
train_file = './hand_head_all_train.txt'
val_file = './hand_head_all_test.txt'

num_classes = 24
train_layers = ['fc8', 'fc7','fc6','conv5','conv4','conv3']
# train_layers = ['fc8','fc7']




optimizers = ['GD']
learning_rates = [0.001]
num_epochses = [10,40,100]
batch_sizes = [16,64]
dropout_rates = [0.1,0.5]

# optimizers = ['Adam','GD']
# learning_rates = [0.1,0.01,0.001,0.0001]
# num_epochses = [10,40,100]
# batch_sizes = [16,32,64]
# dropout_rates = [0.1,0.3,0.5,0.7,1.0]

# s_rowN = 0
# writer_loss = pd.ExcelWriter('./alexnet_multi_choice_loss.xlsx',engine='openpyxl')
# writer_train = pd.ExcelWriter('./alexnet_multi_choice_trainACC.xlsx',engine='openpyxl')
# writer_test = pd.ExcelWriter('./alexnet_multi_choice_testACC.xlsx',engine='openpyxl') 
all_choice_loss = []

for num_epochs in num_epochses:
  for opt in optimizers:
    for learning_rate in learning_rates:
      for batch_size in batch_sizes:
        for dropout_rate in dropout_rates:
          
          tf.reset_default_graph() 

          print('opt = %s , lr = %s , epoch = %d , bt_s = %d , dp = %s ' % (opt, str(learning_rate),num_epochs,batch_size,str(dropout_rate)))


          # How often we want to write the tf.summary data to disk
          display_step = 1

          # Path for tf.summary.FileWriter and to store model checkpoints
          filewriter_path = "C:/Users/GG3BE2/Desktop/winnie/cedl_2017fall/"
          checkpoint_path = "C:/Users/GG3BE2/Desktop/winnie/cedl_2017fall/"

          # Create parent path if it doesn't exist
          if not os.path.isdir(checkpoint_path): os.mkdir(checkpoint_path)



          # TF placeholder for graph input and output
          # concate_score = tf.placeholder(tf.float32, [None, 8192])
          x = tf.placeholder(tf.float32, [batch_size, 227, 227, 3])
          y = tf.placeholder(tf.float32, [None, num_classes])
          head_x = tf.placeholder(tf.float32, [batch_size, 227, 227, 3])
          keep_prob = tf.placeholder(tf.float32)

          # Initialize model
          model = AlexNet(x, keep_prob, num_classes, train_layers)
          model_parall2 = AlexNet(head_x, keep_prob, num_classes, train_layers)


          # Link variable to model output
          score_parall1 = model.fc6
          score_parall2 = model_parall2.fc6
          # model_LastTwo = SimpleNN(tf.concat(1,[score_parall1, score_parall2]), keep_prob, num_classes)
          model_LastTwo = SimpleNN(tf.concat([score_parall1, score_parall2],1), keep_prob, num_classes)


          score = model_LastTwo.fc8 # concate the 2 features (hand and head)
          # score = model.fc8


          # List of trainable variables of the layers we want to train
          var_list = [v for v in tf.trainable_variables() if v.name.split('/')[0] in train_layers]

          # Op for calculating the loss
          with tf.name_scope("cross_ent"):
            loss = tf.reduce_mean(tf.nn.softmax_cross_entropy_with_logits(logits = score, labels = y))  

          # Train op
          with tf.name_scope("train"):
            # Get gradients of all trainable variables
            gradients = tf.gradients(loss, var_list)
            gradients = list(zip(gradients, var_list))
            
            # Create optimizer and apply gradient descent to the trainable variables
            if opt == 'Adam':
              optimizer = tf.train.AdamOptimizer(learning_rate)
            elif opt == 'GD':
              optimizer = tf.train.GradientDescentOptimizer(learning_rate)
            else:
              print('Oops! something wrong!')
            train_op = optimizer.apply_gradients(grads_and_vars=gradients)

          # Add gradients to summary  
          # for gradient, var in gradients:
          #   tf.summary.histogram(var.name + '/gradient', gradient)

          # Add the variables we train to the summary  
          # for var in var_list:
          #   tf.summary.histogram(var.name, var)
            
          # Add the loss to summary
          tf.summary.scalar('cross_entropy', loss)  

          # Evaluation op: Accuracy of the model
          with tf.name_scope("accuracy"):
            print('score = ',score)
            print('score-shape = ',score.get_shape())
            correct_pred = tf.equal(tf.argmax(score, 1), tf.argmax(y, 1))
            accuracy = tf.reduce_mean(tf.cast(correct_pred, tf.float32))

            
          # Add the accuracy to the summary
          tf.summary.scalar('accuracy', accuracy)

          # Merge all summaries together
          merged_summary = tf.summary.merge_all()

          # Initialize the FileWriter
          writer = tf.summary.FileWriter(filewriter_path)

          # Initialize an saver for store model checkpoints
          saver = tf.train.Saver()

          # Initalize the data generator seperately for the training and validation set
          train_generator = ImageDataGenerator(train_file, 
                                               horizontal_flip = True, shuffle = True)
          # train_generator_parall2 = ImageDataGenerator(head_train_file, 
          #                                      horizontal_flip = True, shuffle = True)
          val_generator = ImageDataGenerator(val_file, shuffle = False)
          # val_generator_parall2 = ImageDataGenerator(head_val_file, shuffle = False)  

          # Get the number of training/validation steps per epoch
          train_batches_per_epoch = np.floor(train_generator.data_size / batch_size).astype(np.int16)
          val_batches_per_epoch = np.floor(val_generator.data_size / batch_size).astype(np.int16)

          # Start Tensorflow session
          with tf.Session() as sess:
           
            # sess.run(tf.reset_default_graph())

            # Initialize all variables
            sess.run(tf.global_variables_initializer())
            
            # Add the model graph to TensorBoard
            writer.add_graph(sess.graph)
            
            # Load the pretrained weights into the non-trainable layer
            model.load_initial_weights(sess)
            
            print("{} Start training...".format(datetime.now()))
            print("{} Open Tensorboard at --logdir {}".format(datetime.now(), 
                                                              filewriter_path))
            
            # Loop over number of epochs

            loss_ary = []
            train_acc_ary = []
            test_acc_ary = []
            for epoch in range(num_epochs):
                  
                  print("{} Epoch number: {}".format(datetime.now(), epoch+1))
                  
                  step = 1
                  
                  while step < train_batches_per_epoch: # if num_data = 1000, batch-size = 10, we will have to go through 100 batches(here 100 called train_batches_per_epoch), which called 1 epoch
                      
                      # Get a batch of images and labels
                      batch_xs, batch_head_xs, batch_ys = train_generator.next_batch(batch_size)

                      # And run the training op
                      _, loss_val = sess.run([train_op,loss], feed_dict={x: batch_xs,
                                                    head_x: batch_head_xs, 
                                                    y: batch_ys, 
                                                    keep_prob: dropout_rate})
                      # loss_ary.append(loss_val)
                      # print ('loss = ', loss_val)
                      # Generate summary with the current batch of data and write to file
                      if step%display_step == 0:
                          s = sess.run(merged_summary, feed_dict={x: batch_xs, 
                                                                  head_x: batch_head_xs,
                                                                  y: batch_ys, 
                                                                  keep_prob: 1.})
                                                                  # concate_score: concate_score_ary})
                          writer.add_summary(s, epoch*train_batches_per_epoch + step)
                          
                      step += 1
                  train_accuracy = accuracy.eval(feed_dict={x: batch_xs, 
                                                            head_x: batch_head_xs,
                                                            y: batch_ys,
                                                            keep_prob: 1.}) 
                  print(' %d th epoch, training accuracy %g' % (epoch, train_accuracy))

                  # Validate the model on the entire validation set
                  print("{} Start validation".format(datetime.now()))
                  test_acc = 0.
                  test_count = 0
                  for _ in range(val_batches_per_epoch):
                      batch_tx, batch_head_tx, batch_ty = val_generator.next_batch(batch_size)

                      acc = sess.run(accuracy, feed_dict={x: batch_tx, 
                                                          head_x: batch_head_tx,
                                                          y: batch_ty, 
                                                          keep_prob: 1.})

                      test_acc += acc
                      test_count += 1
                  test_acc /= test_count
                  print("{} Validation Accuracy = {:.4f}".format(datetime.now(), test_acc))
                  
                  # Reset the file pointer of the image data generator
                  val_generator.reset_pointer()
                  train_generator.reset_pointer()
                  
                  print("{} Saving checkpoint of model...".format(datetime.now()))  

                  # train_acc_ary.append(train_accuracy)
                  # test_acc_ary.append(test_acc)
            
            ## loss_ary.append(train_accuracy)
            ## loss_ary.append(test_acc)
            # loss_ary = np.array([s_rowN,3*s_rowN])
            # train_acc_ary = np.array([s_rowN+1,s_rowN+1,s_rowN+1])
            # test_acc_ary = np.array(range(s_rowN))
            '''
            loss_ary = np.asarray([loss_ary])
            train_acc_ary = np.array([train_acc_ary])
            test_acc_ary = np.array([test_acc_ary])
            s_rowN +=1 

            # if isinstance(all_choice_loss, list):
            #   all_choice_loss = loss_ary
            # else:
            #   all_choice_loss = np.concatenate((all_choice_loss,loss_ary),axis=0)
          
            #, engine='openpyxl')
            
            if os.path.isfile('./alexnet_multi_choice_loss.xlsx'):
            #   print('hi, second save')
              print('file already exist, append behind the row.')
              writer_loss.book = load_workbook('./alexnet_multi_choice_loss.xlsx')
              writer_loss.sheets = dict((ws.title, ws) for ws in writer_loss.book.worksheets)
              writer_train.book = load_workbook('./alexnet_multi_choice_trainACC.xlsx')
              writer_train.sheets = dict((ws.title, ws) for ws in writer_train.book.worksheets)
              writer_test.book = load_workbook('./alexnet_multi_choice_testACC.xlsx')
              writer_test.sheets = dict((ws.title, ws) for ws in writer_test.book.worksheets)
            
            print('save loss dataframe')
            df_loss = pd.DataFrame(loss_ary)
            df_loss.to_excel(writer_loss, 'Main' , startrow = s_rowN , header=None)
            writer_loss.save()

            # if os.path.isfile('./alexnet_multi_choice_trainACC.xlsx'):
            #   print('hi, second save')
              
              
            print('save train_acc dataframe')
            df_tr = pd.DataFrame(train_acc_ary)
            df_tr.to_excel(writer_train, 'Main' , startrow = s_rowN , header=None)
            writer_train.save()

            # if os.path.isfile('./alexnet_multi_choice_testACC.xlsx'):
            #   print('hi, second save')
              
              
            
            print('save test_acc dataframe')
            df_ts = pd.DataFrame(test_acc_ary)
            df_ts.to_excel(writer_test, 'Main' , startrow = s_rowN , header=None)
            writer_test.save()

            '''
        
        #save checkpoint of the model
        # checkpoint_name = os.path.join(checkpoint_path, 'model_epoch'+str(epoch+1)+'.ckpt')
        # save_path = saver.save(sess, checkpoint_name)  
        
        # print("{} Model checkpoint saved at {}".format(datetime.now(), checkpoint_name))
        
# writer = pd.ExcelWriter('alexnet_multi_choice_loss.xlsx', engine='xlsxwriter')
# df = pd.DataFrame(all_choice_loss)
# df.to_excel(writer, sheet_name='Sheet1')
# writer.save()


